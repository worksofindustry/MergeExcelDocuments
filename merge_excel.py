# =============================================================================
'''
 __  __                       _____              _ 
|  \/  | ___ _ __ __ _  ___  | ____|_  _____ ___| |
| |\/| |/ _ \ '__/ _` |/ _ \ |  _| \ \/ / __/ _ \ |
| |  | |  __/ | | (_| |  __/ | |___ >  < (_|  __/ |
|_|  |_|\___|_|  \__, |\___| |_____/_/\_\___\___|_|
                 |___/                             
 ____                                        _       
|  _ \  ___   ___ _   _ _ __ ___   ___ _ __ | |_ ___ 
| | | |/ _ \ / __| | | | '_ ` _ \ / _ \ '_ \| __/ __|
| |_| | (_) | (__| |_| | | | | | |  __/ | | | |_\__ \
|____/ \___/ \___|\__,_|_| |_| |_|\___|_| |_|\__|___/

'''
# Merges many Excel docs into one document
# Author: Matt Linker
# =============================================================================
import os
import openpyxl
import numpy as np



### --region Remove Empty Rows from Excel
workbooks = os.listdir('./tmp/')
workbooks = [_ for _ in workbooks if not _.startswith('~')]

if not os.path.exists('tmp'):
    os.makedirs('tmp')
local_path = os.getcwd()
staging_dir = r'tmp'

for workbook in workbooks:
    wb2 = openpyxl.load_workbook('./tmp/' + workbook)
    for sheet in wb2.worksheets:
        max_row_in_sheet = sheet.max_row
        max_col_in_sheet = sheet.max_column
        array_3 = np.array([])
        array_4 = np.array([])
        r = 1
        c = 1
        for r in range(1, max_row_in_sheet+1):
            array_1 = np.array([])
            array_2 = np.array([])
            for c in range (1, max_col_in_sheet+1):
                if sheet.cell(row = r, column = c).value == None:
                    array_1 = np.append(array_2, c)
                    array_2 = array_1
            if len(array_1) == max_col_in_sheet:
                array_3 = np.append(array_4, r)
                array_4 = array_3
                array_3 = array_3.astype(int)
        if len(array_3) != 0:
            index_of_last_array_element = len(array_3) - 1
            while index_of_last_array_element != -1:
                sheet.delete_rows(array_3[index_of_last_array_element], 1)
                index_of_last_array_element = index_of_last_array_element - 1
    save_workbook_to = local_path + r'\\' + staging_dir + r'\\' + workbook
    wb2.save(save_workbook_to)
